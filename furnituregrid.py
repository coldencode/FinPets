import json
import openpyxl
import os
import pandas as pd
import pygame

from create_furniture import create_furniture
from furnitureloader import Furniture
from constants import *

class FurnitureGrid:
    """A class to represent the furniture grid and manage storage/load operations."""
    def __init__(self, id=None, grid_size=GRID_SIZE):
        # Initialize the furniture grid as a 2D list of None values
        self.grid_size = grid_size

        self.excel_path = FURNITURE_GRID_DIRECTORY + str(id) + ".xlsx"
        
        self.furniture_grid = [[None for _ in range(self.grid_size)] for _ in range(self.grid_size)]
        if os.path.exists(self.excel_path):
            self.load_from_excel(self.excel_path)
        
    def add_furniture(self, x, y, furniture):
        """Add a piece of furniture to the grid at the specified location."""
        if 0 <= x < GRID_SIZE and 0 <= y < GRID_SIZE:
            self.furniture_grid[x][y] = furniture
            print(f"Added {furniture} at ({x}, {y})")
        else:
            print("Invalid grid coordinates.")

    def remove_furniture(self, x, y):
        """Remove the furniture from the grid at the specified location."""
        if 0 <= x < GRID_SIZE and 0 <= y < GRID_SIZE:
            removed_furniture = self.furniture_grid[x][y]
            self.furniture_grid[x][y] = None
            print(f"Removed {removed_furniture} from ({x}, {y})")
        else:
            print("Invalid grid coordinates.")

    def save_to_excel(self, filename="furniture_grid.xlsx"):
        """Save the furniture grid to an Excel file."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Furniture Grid"
        for i, row in enumerate(self.furniture_grid):
            for j, furniture in enumerate(row):
                cell = sheet.cell(row=i + 1, column=j + 1)
                #if furniture and furniture != PLACEHOLDER:
                if furniture:
                    if furniture != PLACEHOLDER:
                        cell.value = f"{furniture.name}"  # You can store more details here if needed
                    else:
                        cell.value = PLACEHOLDER
                else:
                    cell.value = None

        # Save the workbook
        workbook.save(filename)
        print(f"Furniture grid saved to {filename}")

    def load_from_excel(self, filename="furniture_grid.xlsx"):
        """Load the furniture grid from an Excel file."""
        df = pd.read_excel("data/furniture.xlsx")
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        for i in range(GRID_SIZE):
            for j in range(GRID_SIZE):
                cell_value = sheet.cell(row=i + 1, column=j + 1).value
                if cell_value:
                    furniture_name = cell_value
                    
                    self.furniture_grid[i][j] = create_furniture(furniture_name)
                else:
                    self.furniture_grid[i][j] = None
        print(f"Furniture grid loaded from {filename}")

    def display_grid(self):
        """Display the grid (for debugging)."""
        for row in self.furniture_grid:
            print([str(furniture) if furniture else "None" for furniture in row])

    def __iter__(self):
        """Make the class iterable by rows."""
        return iter(self.furniture_grid)

    def __getitem__(self, key):
        """Return the item at the specified (x, y) position."""
        x, y = key
        if 0 <= x < GRID_SIZE and 0 <= y < GRID_SIZE:
            return self.furniture_grid[x][y]
        else:
            raise IndexError("Grid index out of range.")

    def __setitem__(self, key, value):
        """Set the item at the specified (x, y) position."""
        x, y = key
        if 0 <= x < GRID_SIZE and 0 <= y < GRID_SIZE:
            self.furniture_grid[x][y] = value
        else:
            raise IndexError("Grid index out of range.")

# Usage example:
if __name__ == "__main__":
    # Initialize the grid and some furniture
    furniture_grid = FurnitureGrid()
    
    # Create some furniture items
    chair = Furniture(1, "Chair", 1, 1)
    sofa = Furniture(2, "Sofa", 2, 1)
    
    # Add furniture to the grid
    furniture_grid.add_furniture(0, 0, chair)
    furniture_grid.add_furniture(1, 0, sofa)
    
    # Display grid (for debugging)
    furniture_grid.display_grid()
    
    # Save the grid to JSON and Excel
    furniture_grid.save_to_excel()

    # Load the grid from JSON and Excel
    furniture_grid.load_from_excel()

    # Display grid again after loading (for debugging)
    furniture_grid.display_grid()